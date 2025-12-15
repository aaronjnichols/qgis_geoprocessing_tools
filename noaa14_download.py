"""
NOAA Atlas 14 Precipitation Frequency Data Tool
Extracts NOAA Atlas 14 data for an area of interest and generates reports.

***************************************************************************
*                                                                         *
*   This program is free software; you can redistribute it and/or modify  *
*   it under the terms of the GNU General Public License as published by  *
*   the Free Software Foundation; either version 2 of the License, or     *
*   (at your option) any later version.                                   *
*                                                                         *
***************************************************************************
"""

from qgis.PyQt.QtCore import QCoreApplication, QVariant
from qgis.core import (
    QgsProcessing,
    QgsProcessingAlgorithm,
    QgsProcessingException,
    QgsProcessingParameterFeatureSource,
    QgsProcessingParameterFileDestination,
    QgsProcessingParameterFolderDestination,
    QgsCoordinateReferenceSystem,
    QgsCoordinateTransform,
    QgsProject,
    QgsFeatureSink,
    QgsProcessingMultiStepFeedback,
    QgsWkbTypes
)
from qgis import processing

import requests
import json
import re
import numpy as np
from pathlib import Path

# Excel/PDF generation imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import ScatterChart, Reference, Series
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class NOAAtlas14ProcessingAlgorithm(QgsProcessingAlgorithm):
    """
    QGIS Processing algorithm to download NOAA Atlas 14 data and create reports.
    """
    
    # Parameter names
    INPUT_LAYER = 'INPUT_LAYER'
    OUTPUT_FOLDER = 'OUTPUT_FOLDER'
    OUTPUT_EXCEL = 'OUTPUT_EXCEL'
    OUTPUT_PDF = 'OUTPUT_PDF'
    
    def tr(self, string):
        """Returns a translatable string with the self.tr() function."""
        return QCoreApplication.translate('Processing', string)
    
    def createInstance(self):
        """Returns a new instance of the algorithm."""
        return NOAAtlas14ProcessingAlgorithm()
    
    def name(self):
        """Returns the algorithm name."""
        return 'noaa_atlas14'
    
    def displayName(self):
        """Returns the translated algorithm name."""
        return self.tr('NOAA Atlas 14 Precipitation Frequency')
    
    def group(self):
        """Returns the name of the group this algorithm belongs to."""
        return self.tr('Data Download Toolbox')
    
    def groupId(self):
        """Returns the unique ID of the group this algorithm belongs to."""
        return 'data_download_toolbox'
    
    def shortHelpString(self):
        """Returns a localised short helper string for the algorithm."""
        return self.tr("""
        Downloads NOAA Atlas 14 precipitation frequency estimates for an area of interest 
        and generates Excel spreadsheet and PDF report with IDF/DDF curves.
        
        <b>Inputs:</b>
        - Area of Interest (polygon layer)
        
        <b>Outputs:</b>
        - Excel spreadsheet with DDF/IDF tables and charts
        - PDF report with precipitation frequency table and curves
        
        <b>Data includes:</b>
        - 19 durations: 5-min through 60-day
        - 10 return periods: 1, 2, 5, 10, 25, 50, 100, 200, 500, 1000 years
        - PDS-based precipitation depths in inches
        
        <b>Requirements:</b>
        - openpyxl (for Excel generation)
        - matplotlib, reportlab (for PDF generation)
        
        The tool calculates the centroid of the input polygon and retrieves 
        precipitation frequency data from NOAA Atlas 14 for that location.
        """)
    
    def initAlgorithm(self, config=None):
        """Define the inputs and outputs of the algorithm."""
        
        # Input: AOI polygon layer
        self.addParameter(
            QgsProcessingParameterFeatureSource(
                self.INPUT_LAYER,
                self.tr('Area of Interest'),
                [QgsProcessing.TypeVectorPolygon]
            )
        )
        
        # Output: Excel file
        self.addParameter(
            QgsProcessingParameterFileDestination(
                self.OUTPUT_EXCEL,
                self.tr('Output Excel File'),
                'Excel Files (*.xlsx)'
            )
        )
        
        # Output: PDF file
        self.addParameter(
            QgsProcessingParameterFileDestination(
                self.OUTPUT_PDF,
                self.tr('Output PDF Report'),
                'PDF Files (*.pdf)'
            )
        )
    
    def processAlgorithm(self, parameters, context, feedback):
        """Execute the algorithm."""
        
        # Check dependencies
        if not OPENPYXL_AVAILABLE:
            raise QgsProcessingException(
                'openpyxl is not installed. Install with: pip install openpyxl'
            )
        
        if not REPORTLAB_AVAILABLE:
            raise QgsProcessingException(
                'matplotlib and/or reportlab are not installed. '
                'Install with: pip install matplotlib reportlab'
            )
        
        # Create multi-step feedback
        multiStepFeedback = QgsProcessingMultiStepFeedback(5, feedback)
        
        # Step 1: Get input layer and calculate centroid
        multiStepFeedback.setCurrentStep(0)
        multiStepFeedback.pushInfo('Calculating centroid of area of interest...')
        
        source = self.parameterAsSource(parameters, self.INPUT_LAYER, context)
        if source is None:
            raise QgsProcessingException(self.invalidSourceError(parameters, self.INPUT_LAYER))
        
        # Get all features and create union
        features = list(source.getFeatures())
        if not features:
            raise QgsProcessingException('Input layer has no features')
        
        # Get geometry union
        from qgis.core import QgsGeometry
        geom_union = QgsGeometry.unaryUnion([f.geometry() for f in features])
        centroid = geom_union.centroid()
        
        # Transform to WGS84 if needed
        source_crs = source.sourceCrs()
        wgs84 = QgsCoordinateReferenceSystem('EPSG:4326')
        
        if source_crs != wgs84:
            transform = QgsCoordinateTransform(source_crs, wgs84, QgsProject.instance())
            centroid.transform(transform)
        
        lat = centroid.asPoint().y()
        lon = centroid.asPoint().x()
        
        multiStepFeedback.pushInfo(f'Centroid: {lat:.6f}°N, {lon:.6f}°W')
        
        # Step 2: Download NOAA Atlas 14 data
        multiStepFeedback.setCurrentStep(1)
        multiStepFeedback.pushInfo('Downloading NOAA Atlas 14 data...')
        
        try:
            noaa_data = self.download_noaa_data(lat, lon, multiStepFeedback)
        except Exception as e:
            raise QgsProcessingException(f'Failed to download NOAA data: {str(e)}')
        
        # Step 3: Create Excel spreadsheet
        multiStepFeedback.setCurrentStep(2)
        multiStepFeedback.pushInfo('Creating Excel spreadsheet...')
        
        excel_path = self.parameterAsFileOutput(parameters, self.OUTPUT_EXCEL, context)
        try:
            self.create_excel_report(noaa_data, excel_path, multiStepFeedback)
            multiStepFeedback.pushInfo(f'Excel file created: {excel_path}')
        except Exception as e:
            raise QgsProcessingException(f'Failed to create Excel file: {str(e)}')
        
        # Step 4: Create PDF report
        multiStepFeedback.setCurrentStep(3)
        multiStepFeedback.pushInfo('Creating PDF report...')
        
        pdf_path = self.parameterAsFileOutput(parameters, self.OUTPUT_PDF, context)
        try:
            self.create_pdf_report(noaa_data, pdf_path, multiStepFeedback)
            multiStepFeedback.pushInfo(f'PDF report created: {pdf_path}')
        except Exception as e:
            raise QgsProcessingException(f'Failed to create PDF report: {str(e)}')
        
        # Step 5: Complete
        multiStepFeedback.setCurrentStep(4)
        multiStepFeedback.pushInfo('Processing complete!')
        multiStepFeedback.pushInfo(f'\nOutputs:')
        multiStepFeedback.pushInfo(f'  Excel: {excel_path}')
        multiStepFeedback.pushInfo(f'  PDF: {pdf_path}')
        
        return {
            self.OUTPUT_EXCEL: excel_path,
            self.OUTPUT_PDF: pdf_path
        }
    
    def download_noaa_data(self, latitude, longitude, feedback):
        """Download NOAA Atlas 14 precipitation frequency data."""
        
        url = "https://hdsc.nws.noaa.gov/cgi-bin/hdsc/new/cgi_readH5.py"
        
        params = {
            'lat': latitude,
            'lon': longitude,
            'type': 'pf',
            'data': 'depth',
            'units': 'english',
            'series': 'pds'
        }
        
        feedback.pushInfo(f'Requesting data from NOAA API...')
        response = requests.get(url, params=params, timeout=30)
        
        if response.status_code != 200:
            raise Exception(f"HTTP {response.status_code}")
        
        # Parse JavaScript arrays from response
        content = response.text
        
        def extract_js_array(content, var_name):
            pattern = rf"{var_name}\s*=\s*(\[\[.*?\]\]);"
            match = re.search(pattern, content, re.DOTALL)
            if match:
                array_str = match.group(1)
                array_str = array_str.replace("'", '"')
                return json.loads(array_str)
            return None
        
        quantiles = extract_js_array(content, 'quantiles')
        upper = extract_js_array(content, 'upper')
        lower = extract_js_array(content, 'lower')
        
        if not quantiles:
            raise Exception("Failed to parse NOAA data from response")
        
        # Standard NOAA Atlas 14 durations
        durations = ['5-min', '10-min', '15-min', '30-min', '60-min', '2-hr', '3-hr', 
                     '6-hr', '12-hr', '24-hr', '2-day', '3-day', '4-day', '7-day', 
                     '10-day', '20-day', '30-day', '45-day', '60-day']
        
        # Standard return periods
        return_periods = [1, 2, 5, 10, 25, 50, 100, 200, 500, 1000]
        
        feedback.pushInfo(f'Retrieved data: {len(durations)} durations, {len(return_periods)} return periods')
        
        return {
            'location': {'latitude': latitude, 'longitude': longitude},
            'durations': durations,
            'return_periods': return_periods,
            'quantiles': quantiles,
            'upper_bounds': upper,
            'lower_bounds': lower
        }
    
    def create_excel_report(self, data, output_file, feedback):
        """Create Excel spreadsheet with DDF/IDF data and charts."""
        
        durations = data['durations']
        return_periods = data['return_periods']
        quantiles = data['quantiles']
        lat = data['location']['latitude']
        lon = data['location']['longitude']
        
        wb = Workbook()
        ws = wb.active
        ws.title = "NOAA Atlas 14 Data"
        
        # Header
        ws['A1'] = 'NOAA Atlas 14 Precipitation Frequency Estimates'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Latitude: {lat:.4f}°, Longitude: {lon:.4f}°"
        ws['A2'].font = Font(size=10)
        ws['A3'] = 'PDS-based depth-duration-frequency (inches)'
        ws['A3'].font = Font(size=10)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # DDF Table
        header_row = 5
        ws.cell(header_row, 1, 'Duration')
        ws.cell(header_row, 1).font = Font(bold=True)
        ws.cell(header_row, 1).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        ws.cell(header_row, 1).border = thin_border
        ws.cell(header_row, 1).alignment = Alignment(horizontal='center', vertical='center')
        
        for i, rp in enumerate(return_periods):
            col = i + 2
            cell = ws.cell(header_row, col, f'{rp}-year')
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for i, duration in enumerate(durations):
            row = header_row + 1 + i
            ws.cell(row, 1, duration)
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 1).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws.cell(row, 1).border = thin_border
            ws.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')
            
            for j, value_str in enumerate(quantiles[i]):
                col = j + 2
                value = float(value_str)
                cell = ws.cell(row, col, value)
                cell.number_format = '0.00'
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.column_dimensions['A'].width = 12
        for i in range(len(return_periods)):
            ws.column_dimensions[chr(66 + i)].width = 10
        
        # Add Charts
        feedback.pushInfo('Adding chart to Excel...')
        
        colors_rp = ['00B050', 'FFC000', 'FF6600', 'FF0000', 'FF00FF', 
                     'C000C0', '0000FF', '0070C0', '00B0F0', '404040']
        
        # Chart 1: DDF Curves
        chart1 = ScatterChart()
        chart1.title = "Depth-Duration-Frequency Curves"
        chart1.style = 2
        chart1.x_axis.title = "Duration"
        chart1.y_axis.title = "Precipitation depth (in)"
        chart1.height = 12
        chart1.width = 18
        chart1.legend.position = 'r'
        
        for rp_idx, rp in enumerate(return_periods):
            xvalues = Reference(ws, min_col=1, min_row=7, max_row=6+len(durations))
            yvalues = Reference(ws, min_col=rp_idx+2, min_row=6, max_row=6+len(durations))
            
            series = Series(yvalues, xvalues, title=f"{rp}-year")
            series.marker.symbol = "none"
            series.graphicalProperties.line.width = 25000
            series.graphicalProperties.line.solidFill = colors_rp[rp_idx]
            series.smooth = True
            
            chart1.series.append(series)
        
        chart1.x_axis.tickLblPos = "low"
        ws.add_chart(chart1, "N5")
        
        wb.save(output_file)
    
    def create_pdf_report(self, data, output_file, feedback):
        """Create PDF report with DDF table and charts."""
        
        durations = data['durations']
        return_periods = data['return_periods']
        quantiles = data['quantiles']
        lat = data['location']['latitude']
        lon = data['location']['longitude']
        
        quantiles_array = np.array([[float(val) for val in row] for row in quantiles])
        
        doc = SimpleDocTemplate(output_file, pagesize=letter,
                               topMargin=0.5*inch, bottomMargin=0.5*inch,
                               leftMargin=0.5*inch, rightMargin=0.5*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=16, textColor=colors.black, spaceAfter=6,
            alignment=TA_CENTER, fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle', parent=styles['Normal'],
            fontSize=10, textColor=colors.black, spaceAfter=12,
            alignment=TA_CENTER
        )
        
        # Page 1: DDF Table
        elements.append(Paragraph("NOAA Atlas 14 Precipitation Frequency Estimates", title_style))
        elements.append(Paragraph(f"Latitude: {lat:.4f}°, Longitude: {lon:.4f}°", subtitle_style))
        elements.append(Paragraph("PDS-based Depth-Duration-Frequency Table (inches)", subtitle_style))
        elements.append(Spacer(1, 0.2*inch))
        
        table_data = [['Duration'] + [f'{rp}-year' for rp in return_periods]]
        for i, duration in enumerate(durations):
            row = [duration] + [f'{float(quantiles[i][j]):.2f}' for j in range(len(return_periods))]
            table_data.append(row)
        
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BACKGROUND', (0, 1), (0, -1), colors.lightgrey),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (0, -1), 8),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTSIZE', (1, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        elements.append(table)
        elements.append(PageBreak())
        
        # Page 2: Charts
        feedback.pushInfo('Creating charts for PDF...')
        
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 10))
        fig.suptitle(f'PDS-based depth-duration-frequency (DDF) curves\n'
                     f'Latitude: {lat:.4f}°, Longitude: {lon:.4f}°', 
                     fontsize=12, fontweight='bold')
        
        colors_rp = ['#00B050', '#FFC000', '#FF6600', '#FF0000', '#FF00FF', 
                     '#C000C0', '#0000FF', '#0070C0', '#00B0F0', '#404040']
        
        x_positions = np.arange(len(durations))
        for rp_idx, rp in enumerate(return_periods):
            depths = quantiles_array[:, rp_idx]
            ax1.plot(x_positions, depths, '-', linewidth=2,
                     color=colors_rp[rp_idx], label=f'{rp}')
        
        ax1.set_xlabel('Duration', fontsize=10)
        ax1.set_ylabel('Precipitation depth (in)', fontsize=10)
        ax1.set_xticks(x_positions)
        ax1.set_xticklabels(durations, rotation=45, ha='right', fontsize=7)
        ax1.grid(True, alpha=0.3)
        ax1.legend(title='Average recurrence\ninterval (years)', 
                  loc='upper left', fontsize=7, title_fontsize=7)
        
        colors_dur = ['#C0C0C0', '#00B050', '#FFC000', '#FF9900', '#FF0000', 
                      '#C00000', '#FF00FF', '#0000FF', '#0070C0', '#00B0F0',
                      '#A0A0A0', '#808080', '#606060', '#404040', '#202020',
                      '#000000', '#000000', '#000000', '#000000']
        
        for dur_idx in range(len(durations)):
            depths = quantiles_array[dur_idx, :]
            ax2.plot(return_periods, depths, '-', linewidth=2,
                     color=colors_dur[dur_idx], label=durations[dur_idx])
        
        ax2.set_xlabel('Average recurrence interval (years)', fontsize=10)
        ax2.set_ylabel('Precipitation depth (in)', fontsize=10)
        ax2.set_xscale('log')
        ax2.set_xlim(1, 1000)
        ax2.grid(True, which='both', alpha=0.3)
        ax2.legend(title='Duration', loc='upper left', fontsize=6, 
                  title_fontsize=7, ncol=2)
        
        plt.tight_layout()
        chart_file = output_file.replace('.pdf', '_charts.png')
        plt.savefig(chart_file, dpi=150, bbox_inches='tight')
        plt.close()
        
        elements.append(Image(chart_file, width=7*inch, height=9*inch))
        
        doc.build(elements)
        
        # Clean up temporary chart file
        import os
        if os.path.exists(chart_file):
            os.remove(chart_file)